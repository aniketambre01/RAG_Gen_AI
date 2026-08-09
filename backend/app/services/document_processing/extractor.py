from pathlib import Path

from pypdf import PdfReader
from docx import Document as DocxDocument


def extract_text(file_path: str) -> str:
    """
    Extract text from a supported document.
    """

    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    extension = path.suffix.lower()

    if extension == ".pdf":
        return extract_pdf(path)

    if extension == ".docx":
        return extract_docx(path)

    if extension in {".txt", ".md", ".py", ".c", ".cpp", ".js", ".ts"}:
        return extract_text_file(path)

    if extension == ".csv":
        return extract_text_file(path)

    if extension == ".xlsx":
        return extract_xlsx(path)

    raise ValueError(
        f"Unsupported document type: {extension}"
    )


def extract_pdf(path: Path) -> str:
    reader = PdfReader(str(path))

    pages = []

    for page in reader.pages:
        text = page.extract_text()

        if text:
            pages.append(text)

    return "\n\n".join(pages)


def extract_docx(path: Path) -> str:
    document = DocxDocument(str(path))

    paragraphs = []

    for paragraph in document.paragraphs:
        if paragraph.text.strip():
            paragraphs.append(paragraph.text)

    return "\n".join(paragraphs)


def extract_text_file(path: Path) -> str:
    return path.read_text(
        encoding="utf-8",
        errors="ignore",
    )


def extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(
        filename=path,
        read_only=True,
        data_only=True,
    )

    rows = []

    for worksheet in workbook.worksheets:
        rows.append(f"Sheet: {worksheet.title}")

        for row in worksheet.iter_rows(values_only=True):
            values = [
                str(value)
                for value in row
                if value is not None
            ]

            if values:
                rows.append(" | ".join(values))

    return "\n".join(rows)