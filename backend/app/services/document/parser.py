from pathlib import Path

from docx import Document
from openpyxl import load_workbook
from pypdf import PdfReader


SUPPORTED_EXTENSIONS = {
    ".pdf",
    ".docx",
    ".xlsx",
    ".csv",
    ".txt",
    ".md",
}


class DocumentParserError(Exception):
    """Raised when a document cannot be parsed."""


def parse_document(file_path: str | Path) -> str:
    """
    Extract text from a supported document.

    Args:
        file_path: Path to the document.

    Returns:
        Extracted text as a string.

    Raises:
        DocumentParserError: If the file type is unsupported
            or parsing fails.
    """
    path = Path(file_path)

    if not path.exists():
        raise DocumentParserError(f"File not found: {path}")

    extension = path.suffix.lower()

    if extension not in SUPPORTED_EXTENSIONS:
        raise DocumentParserError(
            f"Unsupported file type: {extension}"
        )

    try:
        if extension == ".pdf":
            return _parse_pdf(path)

        if extension == ".docx":
            return _parse_docx(path)

        if extension == ".xlsx":
            return _parse_xlsx(path)

        if extension in {".csv", ".txt", ".md"}:
            return _parse_text(path)

    except Exception as exc:
        raise DocumentParserError(
            f"Failed to parse '{path.name}': {exc}"
        ) from exc

    raise DocumentParserError(
        f"No parser available for: {extension}"
    )


def _parse_pdf(path: Path) -> str:
    reader = PdfReader(str(path))

    pages = []

    for page in reader.pages:
        text = page.extract_text() or ""

        if text.strip():
            pages.append(text.strip())

    return "\n\n".join(pages)


def _parse_docx(path: Path) -> str:
    document = Document(str(path))

    paragraphs = [
        paragraph.text.strip()
        for paragraph in document.paragraphs
        if paragraph.text.strip()
    ]

    return "\n\n".join(paragraphs)


def _parse_xlsx(path: Path) -> str:
    workbook = load_workbook(
        filename=str(path),
        read_only=True,
        data_only=True,
    )

    sections = []

    for worksheet in workbook.worksheets:
        sections.append(f"Sheet: {worksheet.title}")

        for row in worksheet.iter_rows(values_only=True):
            values = [
                str(value).strip()
                for value in row
                if value is not None
            ]

            if values:
                sections.append(" | ".join(values))

    workbook.close()

    return "\n".join(sections)


def _parse_text(path: Path) -> str:
    return path.read_text(
        encoding="utf-8",
        errors="replace",
    )